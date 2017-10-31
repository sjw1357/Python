#coding:utf-8

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'test.xlsx'

# ws1 = wb.active  # 第一个表
# ws1.title = "range names"  # 第一个表命名

# 遍历第一个表的1到40行，赋值一个600内的随机数
# for row in range(1, 40):
#     ws1.append(range(0, 6))
ws0 = wb.active
ws0.title = 'mysheet'
_ = ws0.cell(column=5, row=3, value="as")


ws1 = wb.create_sheet('nice')
ws1['A3'] = 'a3..'

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(1, 10):
    for col in range(7, 27):
        # _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
        key = get_column_letter(col) + str(row)
        ws3[key] = get_column_letter(col)
#
wb.save(filename=dest_filename)
